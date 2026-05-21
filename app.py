import os
import re
import math
import json
import shutil
import threading
import queue
from gtts import gTTS
import pygame
import tempfile
import winsound
from datetime import datetime
from collections import defaultdict
from time import sleep
import customtkinter as ctk
import tkinter.filedialog as fd
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2.errors import PdfReadError

from google.oauth2.service_account import Credentials
import gspread

CONFIG_FILE = "config.json"

PERMINTAAN_RESTOCK_SHEET_NAME = "PERMINTAAN_RESTOCK"
# Status yg termasuk "WIP" — untuk display info di scanner, eksekusi, sheet Pesanan.
# Catatan v10.x: WIP HANYA info, tidak lagi mengurangi kekurangan produksi.
RESTOCK_WIP_STATUSES = {"pending", "in_progress", "menunggu_approval"}

class BotApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Bot Sortir Stiker & Gudang v10.7")
        self.geometry("850x700")

        self.config_data = self.load_config()

        # Tabs Utama
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(padx=20, pady=20, fill="both", expand=True)

        self.tab1 = self.tabview.add("Koneksi Gudang")
        self.tab2 = self.tabview.add("Pengaturan File")
        self.tab3 = self.tabview.add("Eksekusi & Log")
        self.tab4 = self.tabview.add("Scanner Resi Gudang")
        self.tab5 = self.tabview.add("Cetak Kekurangan")

        self.setup_tab_koneksi()
        self.setup_tab_file()
        self.setup_tab_eksekusi()
        self.setup_tab_scanner()
        self.setup_tab_kekurangan()

        self.gs_client = None
        self.spreadsheet = None

        self.scanner_db = None
        self.scanner_stock = None
        self.scanner_wip = None
        self.speech_queue = queue.Queue()
        
        # Init pygame mixer for TTS
        pygame.mixer.init()
        
        self.tts_thread = threading.Thread(target=self.tts_worker, daemon=True)
        self.tts_thread.start()

    def tts_worker(self):
        while True:
            text = self.speech_queue.get()
            if text is None:
                break
            try:
                tts = gTTS(text=text, lang='id')
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.mp3')
                temp_file.close()
                tts.save(temp_file.name)
                pygame.mixer.music.load(temp_file.name)
                pygame.mixer.music.play()
                while pygame.mixer.music.get_busy():
                    pygame.time.Clock().tick(10)
                pygame.mixer.music.unload()
                os.remove(temp_file.name)
            except Exception as e:
                print("TTS Error:", e)
            finally:
                self.speech_queue.task_done()

    def speak(self, text):
        self.speech_queue.put(text)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def save_config(self):
        with open(CONFIG_FILE, "w") as f:
            json.dump(self.config_data, f, indent=4)

    # --- TAB 1: Koneksi Gudang ---
    def setup_tab_koneksi(self):
        ctk.CTkLabel(self.tab1, text="Pengaturan Google Spreadsheet", font=("Segoe UI", 16, "bold")).pack(pady=10)
        
        self.url_frame = ctk.CTkFrame(self.tab1)
        self.url_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(self.url_frame, text="URL Spreadsheet:").pack(side="left", padx=10)
        self.entry_url = ctk.CTkEntry(self.url_frame, width=400)
        self.entry_url.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_url.insert(0, self.config_data.get("gsheet_url", ""))

        self.json_frame = ctk.CTkFrame(self.tab1)
        self.json_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(self.json_frame, text="File JSON Credentials:").pack(side="left", padx=10)
        self.entry_json = ctk.CTkEntry(self.json_frame, width=300)
        self.entry_json.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_json.insert(0, self.config_data.get("json_path", ""))
        self.btn_browse_json = ctk.CTkButton(self.json_frame, text="Cari JSON", command=self.browse_json)
        self.btn_browse_json.pack(side="left", padx=10)

        self.btn_test_conn = ctk.CTkButton(self.tab1, text="Test & Simpan Koneksi", command=self.test_connection)
        self.btn_test_conn.pack(pady=20)
        
        self.lbl_conn_status = ctk.CTkLabel(self.tab1, text="Status Koneksi: Belum Dites", text_color="gray")
        self.lbl_conn_status.pack()

    def browse_json(self):
        path = fd.askopenfilename(filetypes=[("JSON Files", "*.json")])
        if path:
            self.entry_json.delete(0, 'end')
            self.entry_json.insert(0, path)

    def test_connection(self):
        url = self.entry_url.get()
        jpath = self.entry_json.get()
        self.config_data["gsheet_url"] = url
        self.config_data["json_path"] = jpath
        self.save_config()
        
        if not url or not os.path.exists(jpath):
            self.lbl_conn_status.configure(text="Error: URL kosong atau JSON tidak ditemukan", text_color="red")
            return
            
        try:
            scopes = ["https://www.googleapis.com/auth/spreadsheets"]
            creds = Credentials.from_service_account_file(jpath, scopes=scopes)
            self.gs_client = gspread.authorize(creds)
            
            if "spreadsheets/d/" in url:
                self.spreadsheet = self.gs_client.open_by_url(url)
            else:
                self.spreadsheet = self.gs_client.open_by_key(url)
                
            # Test Read target sheets, make sure they exist
            self.spreadsheet.worksheet("DATABASE_STIKER")
            self.spreadsheet.worksheet("LOG_KELUAR")
                
            self.lbl_conn_status.configure(text="Berhasil Terhubung & Sheet Ditemukan!", text_color="green")
        except Exception as e:
            error_msg = str(e) if str(e) else repr(e)
            self.lbl_conn_status.configure(text=f"Gagal Terhubung: {error_msg[:60]}", text_color="red")

    # --- TAB 2: Pengaturan File ---
    def setup_tab_file(self):
        ctk.CTkLabel(self.tab2, text="Pengaturan Path File", font=("Segoe UI", 16, "bold")).pack(pady=10)
        
        self.excel_frame = ctk.CTkFrame(self.tab2)
        self.excel_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(self.excel_frame, text="Data Pesanan (.xlsx):").pack(side="left", padx=10)
        self.entry_excel = ctk.CTkEntry(self.excel_frame, placeholder_text="contoh: DATA-V10.xlsx (nama unik per versi bot)")
        self.entry_excel.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_excel.insert(0, self.config_data.get("excel_path", ""))
        ctk.CTkButton(self.excel_frame, text="Browse", width=80, command=lambda: self.browse_path("excel")).pack(side="left", padx=10)

        self.master_frame = ctk.CTkFrame(self.tab2)
        self.master_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(self.master_frame, text="Folder Master PDF:").pack(side="left", padx=10)
        self.entry_master = ctk.CTkEntry(self.master_frame)
        self.entry_master.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_master.insert(0, self.config_data.get("master_path", ""))
        ctk.CTkButton(self.master_frame, text="Browse", width=80, command=lambda: self.browse_path("master")).pack(side="left", padx=10)

        self.hot_frame = ctk.CTkFrame(self.tab2)
        self.hot_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(self.hot_frame, text="Hot Folder (Hasil):").pack(side="left", padx=10)
        self.entry_hot = ctk.CTkEntry(self.hot_frame)
        self.entry_hot.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_hot.insert(0, self.config_data.get("hot_path", ""))
        ctk.CTkButton(self.hot_frame, text="Browse", width=80, command=lambda: self.browse_path("hot")).pack(side="left", padx=10)
        
        ctk.CTkButton(self.tab2, text="Simpan Path", command=self.save_paths).pack(pady=20)

    def browse_path(self, ptype):
        if ptype == "excel":
            path = fd.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if path:
                self.entry_excel.delete(0, 'end')
                self.entry_excel.insert(0, path)
        elif ptype == "master":
            path = fd.askdirectory()
            if path:
                self.entry_master.delete(0, 'end')
                self.entry_master.insert(0, path)
        elif ptype == "hot":
            path = fd.askdirectory()
            if path:
                self.entry_hot.delete(0, 'end')
                self.entry_hot.insert(0, path)
        elif ptype == "kekurangan":
            path = fd.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if path:
                self.entry_kekurangan.delete(0, 'end')
                self.entry_kekurangan.insert(0, path)

    def save_paths(self):
        self.config_data["excel_path"] = self.entry_excel.get()
        self.config_data["master_path"] = self.entry_master.get()
        self.config_data["hot_path"] = self.entry_hot.get()
        self.save_config()
        messagebox.showinfo("Sukses", "Path berhasil disimpan!")

    # --- TAB 3: Eksekusi & Log ---
    def setup_tab_eksekusi(self):
        # === Opsi Pra-Eksekusi (di atas tombol MULAI) ===
        self.opt_frame = ctk.CTkFrame(self.tab3, fg_color="#1f2937", border_width=1, border_color="#3b82f6")
        self.opt_frame.pack(fill="x", padx=20, pady=(15, 5))

        ctk.CTkLabel(
            self.opt_frame,
            text="Opsi Pra-Eksekusi",
            font=("Segoe UI", 13, "bold"),
            text_color="#60a5fa"
        ).pack(anchor="w", padx=12, pady=(8, 2))

        self.var_auto_log_keluar = ctk.BooleanVar(value=self.config_data.get("auto_log_keluar", True))
        self.chk_auto_log = ctk.CTkCheckBox(
            self.opt_frame,
            text="Tulis otomatis ke LOG_KELUAR & kurangi stok gudang saat tersedia",
            variable=self.var_auto_log_keluar,
            command=self.on_toggle_auto_log,
            font=("Segoe UI", 12)
        )
        self.chk_auto_log.pack(anchor="w", padx=15, pady=(2, 4))

        self.lbl_opt_hint = ctk.CTkLabel(
            self.opt_frame,
            text="",
            font=("Segoe UI", 10, "italic"),
            text_color="#9ca3af"
        )
        self.lbl_opt_hint.pack(anchor="w", padx=15, pady=(0, 8))
        self._refresh_opt_hint()

        self.btn_frame = ctk.CTkFrame(self.tab3, fg_color="transparent")
        self.btn_frame.pack(pady=10)

        self.btn_start = ctk.CTkButton(self.btn_frame, text="MULAI PROSES", command=self.start_thread, height=40, font=("Segoe UI", 14, "bold"))
        self.btn_start.pack(side="left", padx=10)

        self.btn_open_output = ctk.CTkButton(self.btn_frame, text="Buka Folder Output", command=self.open_output_folder, height=40, font=("Segoe UI", 14, "bold"), fg_color="#6c757d", hover_color="#5a6268")
        self.btn_open_output.pack(side="left", padx=10)

        self.progress = ctk.CTkProgressBar(self.tab3)
        self.progress.pack(fill="x", padx=20, pady=5)
        self.progress.set(0)
        
        # Sub-tab untuk Log Eksekusi Utama dan Daftar Gudang Ready
        self.sub_tabview = ctk.CTkTabview(self.tab3)
        self.sub_tabview.pack(fill="both", expand=True, padx=20, pady=5)
        self.sub_tab_log = self.sub_tabview.add("Log Eksekusi")
        self.sub_tab_gudang = self.sub_tabview.add("Daftar Ready Gudang")

        self.legend_frame = ctk.CTkFrame(self.sub_tab_log, fg_color="transparent")
        self.legend_frame.pack(fill="x", padx=5, pady=(5, 5))
        
        ctk.CTkLabel(self.legend_frame, text="● Ambil Gudang", text_color="#28a745", font=("Segoe UI", 12, "bold")).pack(side="left", padx=10)
        ctk.CTkLabel(self.legend_frame, text="● Cetak Seluruhnya", text_color="#17a2b8", font=("Segoe UI", 12, "bold")).pack(side="left", padx=10)
        ctk.CTkLabel(self.legend_frame, text="● Error/Gagal", text_color="#dc3545", font=("Segoe UI", 12, "bold")).pack(side="left", padx=10)

        # Log Textbox Utama (dengan warna tag_config)
        self.textbox = ctk.CTkTextbox(self.sub_tab_log, state="disabled")
        self.textbox.pack(fill="both", expand=True, padx=5, pady=5)
        self.textbox.tag_config("hijau", foreground="#28a745")    # Mengambil gudang FULL
        self.textbox.tag_config("kuning", foreground="#ffc107")   # Parsial
        self.textbox.tag_config("merah", foreground="#dc3545")    # Error / Gagal
        self.textbox.tag_config("cyan", foreground="#17a2b8")     # Cetak murni
        self.textbox.tag_config("info", foreground="#adb5bd")   # Info Standar

        # Log Textbox Khusus Daftar Gudang
        self.textbox_gudang = ctk.CTkTextbox(self.sub_tab_gudang, state="disabled", text_color="#28a745")
        self.textbox_gudang.pack(fill="both", expand=True, padx=5, pady=5)

    def _refresh_opt_hint(self):
        if self.var_auto_log_keluar.get():
            self.lbl_opt_hint.configure(
                text="Mode AKTIF: stok yang terpenuhi gudang akan dicatat ke sheet LOG_KELUAR (stok berkurang otomatis).",
                text_color="#86efac"
            )
        else:
            self.lbl_opt_hint.configure(
                text="Mode NONAKTIF: tidak menulis ke LOG_KELUAR. Stok gudang HARUS dipotong manual oleh operator.",
                text_color="#fbbf24"
            )

    def on_toggle_auto_log(self):
        self.config_data["auto_log_keluar"] = bool(self.var_auto_log_keluar.get())
        self.save_config()
        self._refresh_opt_hint()

    def log_gui(self, message, color_tag="info"):
        self.textbox.configure(state="normal")
        self.textbox.insert("end", f"{message}\n", color_tag)
        self.textbox.see("end")
        self.textbox.configure(state="disabled")

    def log_gudang_ready(self, message):
        self.textbox_gudang.configure(state="normal")
        self.textbox_gudang.insert("end", f"{message}\n")
        self.textbox_gudang.see("end")
        self.textbox_gudang.configure(state="disabled")

    # --- TAB 4: Scanner Resi Gudang ---
    def setup_tab_scanner(self):
        ctk.CTkLabel(self.tab4, text="Pemindai Barcode Resi", font=("Segoe UI", 16, "bold")).pack(pady=10)
        
        btn_load = ctk.CTkButton(self.tab4, text="Muat/Perbarui Data Scanner", command=self.confirm_load_scanner_data)
        btn_load.pack(pady=5)
        
        self.lbl_scanner_status = ctk.CTkLabel(self.tab4, text="Status: Belum Memuat Data", text_color="orange")
        self.lbl_scanner_status.pack()

        # Input form sangat besar untuk di scan
        self.entry_scan = ctk.CTkEntry(self.tab4, placeholder_text="Arahkan kursor kesini, lalu scan barcode", height=50, font=("Segoe UI", 20))
        self.entry_scan.pack(fill="x", padx=40, pady=20)
        self.entry_scan.bind("<Return>", self.on_scan)

        # Log Textbox untuk hasil scan
        self.log_scan = ctk.CTkTextbox(self.tab4, state="disabled")
        self.log_scan.pack(fill="both", expand=True, padx=20, pady=10)
        self.log_scan.tag_config("info", foreground="#adb5bd")
        self.log_scan.tag_config("hijau", foreground="#28a745")
        self.log_scan.tag_config("merah", foreground="#dc3545")
        self.log_scan.tag_config("kuning", foreground="#ffc107")

    # --- TAB 5: Cetak Kekurangan Produksi ---
    def setup_tab_kekurangan(self):
        ctk.CTkLabel(
            self.tab5,
            text="Cetak Kekurangan Produksi",
            font=("Segoe UI", 16, "bold")
        ).pack(pady=(15, 4))

        info_frame = ctk.CTkFrame(self.tab5, fg_color="#1f2937", border_width=1, border_color="#dc6803")
        info_frame.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkLabel(
            info_frame,
            text=(
                "Mode untuk request reprint dari tim gudang/packing kalau ada kekurangan.\n"
                "Format Excel: kolom A = SKU (angka, mis. 445), kolom B = Jumlah Lembar (mis. 1).\n"
                "TIDAK menyentuh DATABASE_STIKER, LOG_KELUAR, maupun Pesanan/LIST_PESANAN."
            ),
            font=("Segoe UI", 11),
            text_color="#fbbf24",
            justify="left"
        ).pack(anchor="w", padx=12, pady=8)

        # Path Excel
        path_frame = ctk.CTkFrame(self.tab5)
        path_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(path_frame, text="Excel Kekurangan (.xlsx):").pack(side="left", padx=10)
        self.entry_kekurangan = ctk.CTkEntry(
            path_frame,
            placeholder_text="contoh: KEKURANGAN-V10.xlsx"
        )
        self.entry_kekurangan.pack(side="left", padx=10, fill="x", expand=True)
        self.entry_kekurangan.insert(0, self.config_data.get("kekurangan_path", ""))
        ctk.CTkButton(path_frame, text="Browse", width=80, command=lambda: self.browse_path("kekurangan")).pack(side="left", padx=5)
        ctk.CTkButton(path_frame, text="Simpan", width=80, command=self.save_kekurangan_path).pack(side="left", padx=5)

        # Action buttons
        btn_frame = ctk.CTkFrame(self.tab5, fg_color="transparent")
        btn_frame.pack(pady=10)
        self.btn_kekurangan = ctk.CTkButton(
            btn_frame,
            text="MULAI CETAK KEKURANGAN",
            command=self.start_thread_kekurangan,
            height=40,
            font=("Segoe UI", 14, "bold"),
            fg_color="#dc6803",
            hover_color="#b54708"
        )
        self.btn_kekurangan.pack(side="left", padx=10)

        self.btn_open_kekurangan_output = ctk.CTkButton(
            btn_frame,
            text="Buka Folder Output",
            command=self.open_output_folder,
            height=40,
            font=("Segoe UI", 14, "bold"),
            fg_color="#6c757d",
            hover_color="#5a6268"
        )
        self.btn_open_kekurangan_output.pack(side="left", padx=10)

        # Progress
        self.progress_kekurangan = ctk.CTkProgressBar(self.tab5)
        self.progress_kekurangan.pack(fill="x", padx=20, pady=5)
        self.progress_kekurangan.set(0)

        # Log textbox
        self.textbox_kekurangan = ctk.CTkTextbox(self.tab5, state="disabled")
        self.textbox_kekurangan.pack(fill="both", expand=True, padx=20, pady=10)
        self.textbox_kekurangan.tag_config("hijau", foreground="#28a745")
        self.textbox_kekurangan.tag_config("merah", foreground="#dc3545")
        self.textbox_kekurangan.tag_config("kuning", foreground="#ffc107")
        self.textbox_kekurangan.tag_config("info", foreground="#adb5bd")

    def log_kekurangan(self, msg, color="info"):
        self.textbox_kekurangan.configure(state="normal")
        self.textbox_kekurangan.insert("end", f"{msg}\n", color)
        self.textbox_kekurangan.see("end")
        self.textbox_kekurangan.configure(state="disabled")

    def save_kekurangan_path(self):
        self.config_data["kekurangan_path"] = self.entry_kekurangan.get()
        self.save_config()
        messagebox.showinfo("Sukses", "Path Excel kekurangan disimpan.")

    def start_thread_kekurangan(self):
        self.btn_kekurangan.configure(state="disabled")
        self.textbox_kekurangan.configure(state="normal")
        self.textbox_kekurangan.delete("1.0", "end")
        self.textbox_kekurangan.configure(state="disabled")
        self.progress_kekurangan.set(0)
        threading.Thread(target=self.run_process_kekurangan, daemon=True).start()

    def run_process_kekurangan(self):
        self.log_kekurangan("[INFO] Mulai proses Cetak Kekurangan Produksi.", "info")
        self.log_kekurangan("[INFO] Stok DATABASE_STIKER, LOG_KELUAR & Pesanan TIDAK akan disentuh.", "info")
        try:
            self.main_logic_kekurangan()
        except Exception as e:
            self.log_kekurangan(f"[ERROR FATAL] {str(e)}", "merah")
        finally:
            self.btn_kekurangan.configure(state="normal", text="CETAK KEMBALI")

    def main_logic_kekurangan(self):
        hot_folder = self.config_data.get("hot_path", "")
        master_folder = self.config_data.get("master_path", "")
        kekurangan_file = self.entry_kekurangan.get().strip() or self.config_data.get("kekurangan_path", "")

        if not master_folder or not hot_folder:
            self.log_kekurangan("ERROR: Set 'Folder Master PDF' dan 'Hot Folder' di tab Pengaturan File dulu.", "merah")
            return
        if not kekurangan_file:
            self.log_kekurangan("ERROR: Path Excel Kekurangan belum diisi.", "merah")
            return
        if not os.path.exists(kekurangan_file):
            self.log_kekurangan(f"ERROR: File Excel tidak ditemukan: {kekurangan_file}", "merah")
            return
        if not os.path.exists(master_folder):
            self.log_kekurangan(f"ERROR: Master Folder tidak ditemukan: {master_folder}", "merah")
            return
        if not os.path.exists(hot_folder):
            self.log_kekurangan(f"ERROR: Hot Folder tidak ditemukan: {hot_folder}", "merah")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_dir = os.path.join(hot_folder, f"Cetak_Kekurangan_{timestamp}")
        os.makedirs(output_dir, exist_ok=True)
        self.log_kekurangan(f"[INFO] Folder output: {output_dir}", "info")

        # Build PDF cache (sama logika dgn main flow)
        self.log_kekurangan("[INFO] Memuat cache PDF dari Master Folder...", "info")
        file_cache = defaultdict(list)
        try:
            for filename in os.listdir(master_folder):
                if filename.lower().endswith(".pdf"):
                    m = re.match(r'^\d+', filename)
                    if m:
                        file_cache[m.group(0)].append(os.path.join(master_folder, filename))
        except Exception as e:
            self.log_kekurangan(f"ERROR baca Master Folder: {e}", "merah")
            return

        # Read Excel
        try:
            wb = load_workbook(kekurangan_file, data_only=True)
            ws = wb.active
        except Exception as e:
            self.log_kekurangan(f"ERROR baca Excel: {e}", "merah")
            return

        tasks = []
        skipped = []
        non_stiker_stats = {
            "sheet": {"unique": set(), "pesanan": 0, "pcs": 0},
            "gantungan_kunci": {"unique": set(), "pesanan": 0, "pcs": 0},
        }
        for row_idx in range(2, ws.max_row + 1):
            sku_val = ws.cell(row_idx, 1).value
            jml_val = ws.cell(row_idx, 2).value
            if sku_val is None and jml_val is None:
                continue
            if sku_val is None or jml_val is None:
                skipped.append((str(sku_val) if sku_val else "", str(jml_val) if jml_val else "", "Kolom kosong"))
                continue
            sku_str = str(sku_val).strip()
            # Skip SKU produk non-stiker (stiker sheet '-VN-', gantungan kunci 'GK-')
            kategori = self.non_stiker_category(sku_str)
            if kategori is not None:
                try:
                    pcs_for_stat = int(jml_val)
                except (ValueError, TypeError):
                    try:
                        pcs_for_stat = int(float(jml_val))
                    except (ValueError, TypeError):
                        pcs_for_stat = 0
                non_stiker_stats[kategori]["unique"].add(sku_str)
                non_stiker_stats[kategori]["pesanan"] += 1
                non_stiker_stats[kategori]["pcs"] += pcs_for_stat
                skipped.append((sku_str, str(jml_val), "Skip - SKU produk non-stiker (sheet/GK)"))
                continue
            m = re.match(r'^(\d+)', sku_str)
            if not m:
                skipped.append((sku_str, str(jml_val), "SKU tidak diawali angka"))
                continue
            numeric_id = m.group(1)
            try:
                jumlah = int(jml_val)
            except (ValueError, TypeError):
                try:
                    jumlah = int(float(jml_val))
                except (ValueError, TypeError):
                    skipped.append((sku_str, str(jml_val), "Jumlah harus angka"))
                    continue
            if jumlah <= 0:
                skipped.append((sku_str, str(jml_val), "Jumlah harus > 0"))
                continue
            tasks.append({'sku': numeric_id, 'jumlah': jumlah, 'original': sku_str})

        if not tasks:
            self.log_kekurangan("[INFO] Tidak ada baris valid di Excel.", "merah")
            for s in skipped[:20]:
                self.log_kekurangan(f"   skip: {s[0]} | {s[1]} -> {s[2]}", "kuning")
            return

        self.log_kekurangan(f"[INFO] {len(tasks)} baris valid ({len(skipped)} di-skip). Mulai cetak...", "info")

        success_logs = []
        fail_logs = list(skipped)
        total = len(tasks)

        for i, task in enumerate(tasks):
            self.progress_kekurangan.set(i / total)
            sku = task['sku']
            n = task['jumlah']

            found_paths = file_cache.get(sku, [])
            if not found_paths:
                fail_logs.append((sku, str(n), "Master PDF tidak ditemukan"))
                self.log_kekurangan(f"❌ SKU {sku}: Master PDF tidak ada", "merah")
                continue

            optimal = [p for p in found_paths if 'versioptimal' in os.path.basename(p).lower()]
            standard = [p for p in found_paths if p not in optimal]
            if optimal:
                optimal.sort(key=len)
                src, ver = optimal[0], "optimal"
            else:
                standard.sort(key=len)
                src, ver = standard[0], "standard"

            try:
                with open(src, "rb") as infile:
                    reader = PdfReader(infile)
                    if len(reader.pages) == 0:
                        fail_logs.append((sku, str(n), "Master PDF kosong"))
                        self.log_kekurangan(f"❌ SKU {sku}: Master PDF kosong", "merah")
                        continue
                    page0 = reader.pages[0]
                    for j in range(1, n + 1):
                        out_name = f"{sku}-{j}.pdf" if n > 1 else f"{sku}.pdf"
                        out_path = os.path.join(output_dir, out_name)
                        writer = PdfWriter()
                        writer.add_page(page0)
                        with open(out_path, "wb") as outfile:
                            writer.write(outfile)
                self.log_kekurangan(f"✔ SKU {sku}: {n} lembar (versi {ver})", "hijau")
                success_logs.append((sku, n, f"versi {ver}"))
            except Exception as e:
                fail_logs.append((sku, str(n), f"Error ekstrak: {e}"))
                self.log_kekurangan(f"❌ SKU {sku}: {e}", "merah")

        self.progress_kekurangan.set(1.0)

        # Save log Excels
        log_dir = os.path.join(output_dir, "log")
        os.makedirs(log_dir, exist_ok=True)
        if success_logs:
            wb_s = Workbook()
            ws_s = wb_s.active
            ws_s.append(["SKU", "Jumlah Lembar", "Keterangan"])
            for r in success_logs:
                ws_s.append(r)
            wb_s.save(os.path.join(log_dir, "berhasil.xlsx"))
        if fail_logs:
            wb_f = Workbook()
            ws_f = wb_f.active
            ws_f.append(["SKU", "Jumlah", "Keterangan"])
            for r in fail_logs:
                ws_f.append(r)
            wb_f.save(os.path.join(log_dir, "gagal.xlsx"))

        # Summary SKU produk non-stiker yg di-skip (kalau ada)
        sheet_stats = non_stiker_stats["sheet"]
        gk_stats = non_stiker_stats["gantungan_kunci"]
        if sheet_stats["pesanan"] > 0 or gk_stats["pesanan"] > 0:
            self.log_kekurangan("\n[INFO] Produk non-stiker yang di-skip (tidak dicetak):", "kuning")
            if sheet_stats["pesanan"] > 0:
                self.log_kekurangan(
                    f"  • Stiker SHEET: {len(sheet_stats['unique'])} SKU unik, "
                    f"{sheet_stats['pesanan']} baris ({sheet_stats['pcs']} pcs total)",
                    "kuning"
                )
            if gk_stats["pesanan"] > 0:
                self.log_kekurangan(
                    f"  • Gantungan KUNCI: {len(gk_stats['unique'])} SKU unik, "
                    f"{gk_stats['pesanan']} baris ({gk_stats['pcs']} pcs total)",
                    "kuning"
                )

        self.log_kekurangan(f"\n[SELESAI] {len(success_logs)} SKU sukses, {len(fail_logs)} gagal/skip.", "info")
        self.log_kekurangan(f"[INFO] Output disimpan di: {output_dir}", "info")

    def _ensure_gs_connected(self):
        if self.spreadsheet:
            return True
        self.test_connection()
        return self.spreadsheet is not None

    def load_wip_map(self):
        """Return dict {sku: total_pcs_in_pipeline} dari PERMINTAAN_RESTOCK.
        Hitung row dgn status in [pending, in_progress, menunggu_approval].
        Pakai Jumlah_Aktual_Gudang (kol I, idx 8) kalau sudah diisi, else Jumlah_Request (kol C, idx 2).
        Layout v10.3: status di kol F (idx 5)."""
        wip_map = {}
        if not self._ensure_gs_connected():
            return wip_map
        try:
            ws = self.spreadsheet.worksheet(PERMINTAAN_RESTOCK_SHEET_NAME)
        except gspread.WorksheetNotFound:
            return wip_map
        try:
            rows = ws.get_all_values()
        except Exception:
            return wip_map
        for r in rows[1:]:
            if len(r) < 6:
                continue
            status = str(r[5]).strip().lower()
            if status not in RESTOCK_WIP_STATUSES:
                continue
            sku = str(r[1]).strip() if len(r) > 1 else ""
            if not sku:
                continue
            pcs_aktual_str = str(r[8]).strip() if len(r) > 8 else ""
            pcs_req_str = str(r[2]).strip() if len(r) > 2 else ""
            pcs = 0
            try:
                pcs = int(pcs_aktual_str) if pcs_aktual_str else 0
            except ValueError:
                pcs = 0
            if pcs <= 0:
                try:
                    pcs = int(pcs_req_str)
                except (ValueError, AttributeError):
                    pcs = 0
            if pcs > 0:
                wip_map[sku] = wip_map.get(sku, 0) + pcs
        return wip_map

    def print_scan_log(self, msg, color="info"):
        self.log_scan.configure(state="normal")
        self.log_scan.insert("end", f"{msg}\n", color)
        self.log_scan.see("end")
        self.log_scan.configure(state="disabled")

    def confirm_load_scanner_data(self):
        answer = messagebox.askyesno("Pengingat Update Sales", "Apakah Anda sudah memastikan 'Upload Data Sales' telah di-update di Google Sheet secara manual sebelum memuat data scanner?")
        if answer:
            threading.Thread(target=self.load_scanner_data, daemon=True).start()

    def load_scanner_data(self):
        self.lbl_scanner_status.configure(text="Status: Sedang memuat data... Mohon tunggu", text_color="orange")
        self.print_scan_log("\n[INFO] Memulai sinkronisasi data gudang & pesanan...", "info")
        
        # Test Connection 
        if not self.spreadsheet:
            self.test_connection()
            if not self.spreadsheet:
                self.print_scan_log("ERROR: Gagal terhubung ke Google Sheets.", "merah")
                self.lbl_scanner_status.configure(text="Status: Gagal Terhubung", text_color="red")
                return
                
        excel_file = self.config_data.get("excel_path", "")
        if not excel_file or not os.path.exists(excel_file):
            self.print_scan_log("ERROR: File Pesanan Excel tidak valid di tab Pengaturan File.", "merah")
            self.lbl_scanner_status.configure(text="Status: Excel tidak valid", text_color="red")
            return

        try:
            # Load GS
            ws_db = self.spreadsheet.worksheet("DATABASE_STIKER")
            db_records = ws_db.get_all_values()
            stock_dict = {}
            for row in db_records[1:]:
                if len(row) >= 8:
                    sku_id = str(row[0]).strip()
                    try: stok_saat_ini = int(row[7])
                    except ValueError: stok_saat_ini = 0
                    stock_dict[sku_id] = stok_saat_ini
            self.scanner_stock = stock_dict

            # Load Excel
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            scan_db = defaultdict(list)
            for row_idx in range(2, ws.max_row + 1):
                resi_val = ws.cell(row_idx, 1).value
                sku_val = ws.cell(row_idx, 2).value
                jml_val = ws.cell(row_idx, 3).value
                if not sku_val or not jml_val or not resi_val: continue
                
                try: jumlah_pesanan = int(jml_val)
                except ValueError: continue
                
                original_sku = str(sku_val).strip()
                # Skip SKU produk non-stiker (stiker sheet '-VN-', gantungan kunci 'GK-')
                if self.is_non_stiker_sku(original_sku):
                    continue
                numeric_id, pcs_per_paket = self.extract_numeric_id_and_pcs(original_sku)
                if not numeric_id: continue

                total_pcs_needed = pcs_per_paket * jumlah_pesanan
                scan_db[str(resi_val).strip()].append({
                    "sku": original_sku,
                    "numeric_id": numeric_id,
                    "total_pcs_needed": total_pcs_needed
                })

            self.scanner_db = scan_db

            # Load WIP map (info-only — tidak dihitung sebagai stok available)
            self.scanner_wip = self.load_wip_map()

            self.print_scan_log(
                f"Berhasil! Dimuat {len(self.scanner_db)} Resi Unik, "
                f"{sum(self.scanner_wip.values())} pcs WIP pending ({len(self.scanner_wip)} SKU).",
                "hijau"
            )
            self.lbl_scanner_status.configure(text="Status: Siap! Arahkan kursor dan Scan Barcode.", text_color="green")
            self.speak("Data siap digunakan")
        except Exception as e:
            self.print_scan_log(f"ERROR: {e}", "merah")
            self.lbl_scanner_status.configure(text="Status: Error Gagal Muat Data", text_color="red")

    def on_scan(self, event):
        resi = self.entry_scan.get().strip()
        self.entry_scan.delete(0, 'end') # Clear
        
        if not resi:
            return
            
        try:
            winsound.Beep(1500, 150)
        except:
            pass
            
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.print_scan_log(f"\n[{timestamp}] > Barcode Terbaca: {resi}", "info")
        
        if self.scanner_db is None or self.scanner_stock is None:
            self.print_scan_log("Proses ditolak. Harap 'Muat/Perbarui Data Scanner' terlebih dahulu.", "merah")
            self.speak("Harap muat data terlebih dahulu")
            return

        if resi not in self.scanner_db:
            self.print_scan_log(f"Resi '{resi}' tidak ditemukan pada file Excel pesanan.", "merah")
            self.speak("Resi tidak ditemukan")
            return

        items = self.scanner_db[resi]
        self.print_scan_log(f"Ditemukan {len(items)} SKU dalam resi ini.", "info")
        
        ready_count = sum(1 for item in items if self.scanner_stock.get(item['numeric_id'], 0) >= item['total_pcs_needed'])
        
        if ready_count == 0:
            self.speak("Semua kosong")
        
        wip_map = getattr(self, 'scanner_wip', {}) or {}

        for item in items:
            num_id = item['numeric_id']
            original_sku = item['sku']
            needed = item['total_pcs_needed']
            current_gudang = self.scanner_stock.get(num_id, 0)
            wip = wip_map.get(num_id, 0)

            if current_gudang >= needed:
                self.print_scan_log(f"✔ SKU {num_id} ({needed} pcs) MENCUKUPI (Stok: {current_gudang})", "hijau")
                self.speak(f"{num_id} ready")
            else:
                sisa_produksi = needed
                msg = f"SKU {num_id} ({needed} pcs) STOK TIDAK CUKUP (Stok: {current_gudang}, Cetak Full: {sisa_produksi})"
                self.print_scan_log(f"❌ {msg}", "merah")
                if ready_count > 0:
                    self.speak(f"{num_id} kosong")

            # Info WIP (kalau ada) — info-only, JANGAN tunda cetak. Tim print
            # prioritas orderan; restock dikerjakan setelah orderan menurun.
            if wip > 0 and current_gudang < needed:
                self.print_scan_log(
                    f"   ℹ WIP info: {wip} pcs restock SKU {num_id} sedang dalam pipeline "
                    f"(tetap cetak orderan — WIP info aja, bukan pengganti stok).",
                    "kuning"
                )

    def open_output_folder(self):
        hot_folder = self.config_data.get("hot_path", "")
        if hot_folder and os.path.exists(hot_folder):
            os.startfile(hot_folder)
        else:
            messagebox.showwarning("Peringatan", "Folder output belum diatur atau tidak ditemukan.")

    def start_thread(self):
        self.btn_start.configure(state="disabled")
        
        # Bersihkan textbox
        self.textbox.configure(state="normal")
        self.textbox.delete("1.0", "end")
        self.textbox.configure(state="disabled")
        
        self.textbox_gudang.configure(state="normal")
        self.textbox_gudang.delete("1.0", "end")
        self.textbox_gudang.configure(state="disabled")
        
        self.progress.set(0)
        
        # Test Connection 
        if not self.spreadsheet:
            self.test_connection()
            if not self.spreadsheet:
                self.log_gui("ERROR: Gagal terhubung ke Google Sheets. Cek pengaturan.", "merah")
                self.btn_start.configure(state="normal")
                return

        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        self.log_gui("[INFO] Memulai Proses...", "info")
        try:
            self.main_logic()
        except Exception as e:
            self.log_gui(f"[ERROR FATAL] Terjadi kesalahan: {str(e)}", "merah")
        finally:
            self.btn_start.configure(state="normal", text="MULAI KEMBALI")

    def clear_hotfolder(self, hot_folder):
        self.log_gui("[INFO] Membersihkan HOT_FOLDER dari file lama...", "info")
        try:
            for item in os.listdir(hot_folder):
                full_path = os.path.join(hot_folder, item)
                if os.path.isdir(full_path):
                    if item.lower() == "log":
                        continue
                    # Hapus folder Batch sebelumnya jika ada
                    if item.lower().startswith("batch"):
                        shutil.rmtree(full_path, ignore_errors=True)
                elif os.path.isfile(full_path) and item.lower().endswith('.pdf'):
                    os.remove(full_path)
            self.log_gui("Pembersihan selesai.", "info")
        except Exception as e:
            self.log_gui(f"Error membersihkan HOT_FOLDER: {e}", "merah")

    def create_file_cache(self, master_folder):
        self.log_gui("[INFO] Membuat cache daftar file dari MASTER_FOLDER...", "info")
        cache = defaultdict(list)
        try:
            for filename in os.listdir(master_folder):
                if filename.lower().endswith(".pdf"):
                    match = re.match(r'^\d+', filename)
                    if match:
                        name_key = match.group(0)
                        cache[name_key].append(os.path.join(master_folder, filename))
            return cache
        except FileNotFoundError:
            self.log_gui(f"Error: MASTER_FOLDER tidak ditemukan.", "merah")
            return None

    @staticmethod
    def is_non_stiker_sku(sku):
        """Return True kalau SKU bukan stiker reguler (harus di-skip dari duplicate/cetak).

        Pola produk lain yang kami tidak proses:
        - Stiker SHEET → mengandung '-VN-' (cth: '136-VN-A6-A'). Tidak ada multiplier pcs,
          jangan masuk pipeline cetak.
        - Gantungan KUNCI → prefix 'GK-' (cth: 'GK-ATM-0010752-L').
        """
        return BotApp.non_stiker_category(sku) is not None

    @staticmethod
    def non_stiker_category(sku):
        """Return 'sheet' kalau SKU = stiker sheet '-VN-', 'gantungan_kunci' kalau prefix
        'GK-', else None (= stiker reguler, boleh diproses)."""
        if not sku:
            return None
        s = str(sku).strip().upper()
        if s.startswith("GK-"):
            return "gantungan_kunci"
        if "-VN-" in s:
            return "sheet"
        return None

    def extract_numeric_id_and_pcs(self, sku):
        id_match = re.match(r'^\d+', sku.strip())
        numeric_id = id_match.group(0) if id_match else None
        pcs_match = re.search(r'(\d+)pcs', sku, re.IGNORECASE)
        pcs = int(pcs_match.group(1)) if pcs_match else 1
        return numeric_id, pcs

    def main_logic(self):
        hot_folder = self.config_data.get("hot_path", "")
        master_folder = self.config_data.get("master_path", "")
        excel_file = self.config_data.get("excel_path", "")

        if not all([hot_folder, master_folder, excel_file]):
            self.log_gui("ERROR: Semua path (Excel, Master, Hot) harus diisi.", "merah")
            return

        if not os.path.exists(excel_file):
            self.log_gui(f"ERROR: File Excel tidak ditemukan: {excel_file}", "merah")
            return

        self.clear_hotfolder(hot_folder)

        # Build paths untuk LOG lokal
        LOG_BERHASIL_DIR = os.path.join(hot_folder, "log", "berhasil")
        LOG_GAGAL_DIR = os.path.join(hot_folder, "log", "gagal")
        LOG_PERINGATAN_DIR = os.path.join(hot_folder, "log", "peringatan")
        os.makedirs(LOG_BERHASIL_DIR, exist_ok=True)
        os.makedirs(LOG_GAGAL_DIR, exist_ok=True)
        os.makedirs(LOG_PERINGATAN_DIR, exist_ok=True)

        file_cache = self.create_file_cache(master_folder)
        if file_cache is None: return

        # Load Stock from GS
        self.log_gui("[INFO] Mengunduh data dari DATABASE_STIKER...", "info")
        ws_db = self.spreadsheet.worksheet("DATABASE_STIKER")
        db_records = ws_db.get_all_values()
        
        stock_dict = {}
        for row in db_records[1:]:
            if len(row) >= 8:
                sku_id = str(row[0]).strip()
                try: stok_saat_ini = int(row[7])
                except ValueError: stok_saat_ini = 0
                stock_dict[sku_id] = stok_saat_ini

        ws_log = self.spreadsheet.worksheet("LOG_KELUAR")

        # Compile Excel rows into Task List
        self.log_gui("[INFO] Membaca & Mengurutkan pesanan Excel (Sort by SKU)...", "info")
        wb = load_workbook(excel_file)
        ws = wb.active
        
        task_list = []
        fail_logs = []
        # Counter SKU produk non-stiker yg di-skip (untuk summary di akhir proses).
        # {category: {"unique": set(sku), "pesanan": total_baris, "pcs": total_qty}}
        non_stiker_stats = {
            "sheet": {"unique": set(), "pesanan": 0, "pcs": 0},
            "gantungan_kunci": {"unique": set(), "pesanan": 0, "pcs": 0},
        }

        for row_idx in range(2, ws.max_row + 1):
            resi_val = ws.cell(row_idx, 1).value
            sku_val = ws.cell(row_idx, 2).value
            jml_val = ws.cell(row_idx, 3).value

            if not sku_val or not jml_val: continue

            try:
                jumlah_pesanan = int(jml_val)
                original_sku = str(sku_val).strip()
            except ValueError:
                fail_logs.append((str(sku_val), str(jml_val), f"Gagal - 'Jumlah' harus angka"))
                continue

            # Skip SKU produk non-stiker (stiker sheet '-VN-', gantungan kunci 'GK-')
            kategori = self.non_stiker_category(original_sku)
            if kategori is not None:
                non_stiker_stats[kategori]["unique"].add(original_sku)
                non_stiker_stats[kategori]["pesanan"] += 1
                non_stiker_stats[kategori]["pcs"] += jumlah_pesanan
                fail_logs.append((original_sku, str(jml_val), "Skip - SKU produk non-stiker (sheet/gantungan kunci), tidak dicetak."))
                continue

            numeric_id, pcs_per_paket = self.extract_numeric_id_and_pcs(original_sku)
            if not numeric_id:
                fail_logs.append((original_sku, str(jml_val), "Gagal - Tidak ada ID (angka awal) terdeteksi."))
                continue

            total_pcs_needed = pcs_per_paket * jumlah_pesanan

            task_list.append({
                'resi': str(resi_val) if resi_val else "-",
                'sku': original_sku,
                'numeric_id': numeric_id,
                'total_pcs_needed': total_pcs_needed,
                'jumlah_pesanan': jumlah_pesanan,
                'pcs_per_paket': pcs_per_paket
            })

        if not task_list:
            self.log_gui("Data kosong atau tidak valid.", "merah")
            return

        # SORTING by Numeric ID
        task_list.sort(key=lambda x: int(x['numeric_id']))

        success_logs = []
        warnings_log = []
        logs_keluar_to_append = []
        auto_log_keluar = bool(self.config_data.get("auto_log_keluar", True))
        if auto_log_keluar:
            self.log_gui("[OPSI] LOG_KELUAR: AKTIF - stok gudang akan dikurangi otomatis.", "info")
        else:
            self.log_gui("[OPSI] LOG_KELUAR: NONAKTIF - stok TIDAK ditulis ke sheet, potong manual.", "kuning")

        total_tasks = len(task_list)
        today_str = datetime.now().strftime("%Y-%m-%d")

        used_filenames = defaultdict(int)
        def get_next_filename(task_id: int, design_id: str, suffix: str = "") -> str:
            sanitized_id = re.sub(r'[\\/*?:"<>|]', "-", design_id)
            base_key = f"{sanitized_id}{suffix}"
            used_filenames[base_key] += 1
            if used_filenames[base_key] == 1: return f"{base_key}.pdf"
            else: return f"{base_key} - Copy ({used_filenames[base_key] - 1}).pdf"

        batches = {
            '10pcs': {'base_dir': os.path.join(hot_folder, "Batch 10"), 'number': 1, 'count': 0, 'dir': os.path.join(hot_folder, "Batch 10", "Batch_1")},
            '50pcs': {'base_dir': os.path.join(hot_folder, "Batch 50"), 'number': 1, 'count': 0, 'dir': os.path.join(hot_folder, "Batch 50", "Batch_1")}
        }
        os.makedirs(batches['10pcs']['dir'], exist_ok=True)
        os.makedirs(batches['50pcs']['dir'], exist_ok=True)

        # Agregat sisa produksi per (numeric_id, batch_type) supaya 10 resi
        # @10pcs (=100pcs total) hanya cetak 1 lembar (100pcs/page optimal),
        # bukan 10 lembar seperti versi lama yang per-baris.
        sku_print_queue = {}  # key: (numeric_id, '10pcs'|'50pcs'), val: dict
        # Map untuk update keterangan di success_logs setelah aggregate print
        # selesai. Index = posisi di success_logs (placeholder PENDING_PRODUKSI).
        pending_success_idx = {}  # (numeric_id, batch_type) -> list of (idx, sisa)

        for i, task in enumerate(task_list):
            self.progress.set((i) / total_tasks)

            numeric_id = task['numeric_id']
            resi_val = task['resi']
            total_pcs_needed = task['total_pcs_needed']
            original_sku = task['sku']

            stok_gudang = stock_dict.get(numeric_id, 0)

            if stok_gudang >= total_pcs_needed:
                # 1. Full terpenuhi gudang
                ambil_terpenuhi = total_pcs_needed
                sisa_produksi = 0
                stock_dict[numeric_id] -= ambil_terpenuhi

                if auto_log_keluar:
                    logs_keluar_to_append.append([today_str, numeric_id, ambil_terpenuhi, f"Resi: {resi_val} | Full"])
                    ket_sukses = f"Tersedia dari gudang ({ambil_terpenuhi})"
                else:
                    ket_sukses = f"Tersedia dari gudang ({ambil_terpenuhi}) - LOG_KELUAR dilewati (manual)"

                self.log_gui(f"● Resi {resi_val} (SKU {numeric_id}): {total_pcs_needed} pcs", "hijau")
                self.log_gudang_ready(f"GUDANG - Resi: {resi_val} | SKU: {numeric_id} | Jumlah: {ambil_terpenuhi} pcs")
                success_logs.append((f"Tugas-{(i+1):03d}", numeric_id, total_pcs_needed, ket_sukses))

            else:
                # 2. Produksi — queue ke print agregat per SKU di akhir.
                sisa_produksi = total_pcs_needed
                self.log_gui(f"● Resi {resi_val} (SKU {numeric_id}): {total_pcs_needed} pcs", "cyan")

                batch_type = '10pcs' if task['pcs_per_paket'] <= 20 else '50pcs'
                key = (numeric_id, batch_type)
                if key not in sku_print_queue:
                    sku_print_queue[key] = {
                        'sisa': 0,
                        'first_task_index': i + 1,
                        'sample_task': task,
                        'task_count': 0,
                    }
                sku_print_queue[key]['sisa'] += sisa_produksi
                sku_print_queue[key]['task_count'] += 1

                # Placeholder di success_logs — akan di-update setelah print.
                placeholder_idx = len(success_logs)
                success_logs.append((f"Tugas-{(i+1):03d}", numeric_id, sisa_produksi, "PENDING_PRODUKSI"))
                pending_success_idx.setdefault(key, []).append((placeholder_idx, sisa_produksi))

        # =====================================================================
        # CETAK AGREGAT PER SKU — hitung jumlah lembar dari TOTAL sisa per SKU,
        # bukan dari sisa per-baris. Ini menghilangkan duplikasi lembar untuk
        # banyak resi kecil dengan SKU sama.
        # =====================================================================
        if sku_print_queue:
            self.log_gui(
                f"\n[INFO] Cetak agregat untuk {len(sku_print_queue)} grup SKU yang butuh produksi...",
                "info"
            )

        # Load WIP map sekali — info-only. Tim print prioritas orderan,
        # cetak tetap jalan walaupun ada WIP pipeline (restock dikerjakan
        # belakangan saat orderan menurun).
        wip_map_main = {}
        try:
            wip_map_main = self.load_wip_map()
            if wip_map_main:
                self.log_gui(
                    f"[INFO] WIP pipeline: {sum(wip_map_main.values())} pcs di {len(wip_map_main)} SKU "
                    f"(sheet PERMINTAAN_RESTOCK) — info aja, cetak orderan tetap jalan.",
                    "info"
                )
        except Exception as e:
            self.log_gui(f"[WARN] Gagal load WIP map: {e}", "kuning")

        for key, q in sku_print_queue.items():
            numeric_id, batch_type = key
            total_sisa = q['sisa']
            task = q['sample_task']
            original_sku = task['sku']

            # Info WIP (tidak menahan cetak). Kekurangan tetap diproses penuh.
            wip_existing = wip_map_main.get(numeric_id, 0)
            if wip_existing > 0:
                self.log_gui(
                    f"   ℹ SKU {numeric_id}: ada {wip_existing} pcs di pipeline restock — "
                    f"tetap cetak {total_sisa} pcs untuk orderan ini (WIP info-only).",
                    "kuning"
                )

            found_paths = file_cache.get(numeric_id, [])
            if not found_paths:
                fail_logs.append((original_sku, f"Sisa Produksi: {total_sisa}", "Gagal - Master PDF tidak ada."))
                self.log_gui(f"❌ (SKU {numeric_id}) File Master tidak ditemukan!", "merah")
                # Update placeholder ke status gagal
                for idx, sisa in pending_success_idx.get(key, []):
                    entry = success_logs[idx]
                    success_logs[idx] = (entry[0], entry[1], entry[2], "Gagal - Master PDF tidak ada")
                continue

            optimal_candidates = [p for p in found_paths if "versioptimal" in os.path.basename(p).lower()]
            standard_candidates = [p for p in found_paths if p not in optimal_candidates]

            found_file = None
            version_type = "standard"
            if optimal_candidates:
                optimal_candidates.sort(key=len)
                found_file = optimal_candidates[0]
                version_type = "optimal"
                if len(optimal_candidates) > 1:
                    warnings_log.append((numeric_id, "", "Duplikat versi optimal ditemukan."))
            elif standard_candidates:
                standard_candidates.sort(key=len)
                found_file = standard_candidates[0]
                if len(standard_candidates) > 1:
                    warnings_log.append((numeric_id, "", "Duplikat standar ditemukan."))

            batch_size_per_page = 100.0 if version_type == 'optimal' else 50.0
            num_pages_to_print = int(math.ceil(total_sisa / batch_size_per_page))

            batch_info = batches[batch_type]
            if batch_info['count'] >= 20:
                batch_info['number'] += 1
                batch_info['count'] = 0
                batch_info['dir'] = os.path.join(batch_info['base_dir'], f"Batch_{batch_info['number']}")
                os.makedirs(batch_info['dir'], exist_ok=True)

            try:
                for _ in range(num_pages_to_print):
                    out_filename = get_next_filename(
                        q['first_task_index'],
                        numeric_id,
                        "-PRODUK" if version_type != 'optimal' else "-VERSIOPTIMAL"
                    )
                    out_path = os.path.join(batch_info['dir'], out_filename)
                    with open(found_file, "rb") as infile:
                        reader = PdfReader(infile)
                        writer = PdfWriter()
                        if len(reader.pages) > 0:
                            writer.add_page(reader.pages[0])
                            with open(out_path, "wb") as outfile:
                                writer.write(outfile)

                ket_agregat = (
                    f"Sukses Cetak {num_pages_to_print} lbr "
                    f"(agregat {total_sisa} pcs dari {q['task_count']} resi → "
                    f"{batch_type}/Batch_{batch_info['number']})"
                )
                self.log_gui(
                    f"➤ SKU {numeric_id}: cetak {num_pages_to_print} lbr "
                    f"(total {total_sisa} pcs dari {q['task_count']} resi → "
                    f"{batch_type}/Batch_{batch_info['number']})",
                    "cyan"
                )
                # Update semua placeholder PENDING_PRODUKSI untuk key ini
                for idx, sisa in pending_success_idx.get(key, []):
                    entry = success_logs[idx]
                    success_logs[idx] = (entry[0], entry[1], entry[2], ket_agregat)

                batch_info['count'] += 1
            except Exception as e:
                fail_logs.append((original_sku, str(total_sisa), f"Gagal Ekstrak - {e}"))
                self.log_gui(f"❌ Error ekstrak SKU {numeric_id}: {e}", "merah")
                for idx, sisa in pending_success_idx.get(key, []):
                    entry = success_logs[idx]
                    success_logs[idx] = (entry[0], entry[1], entry[2], f"Gagal Ekstrak - {e}")

        # PUSH SHEETS
        if logs_keluar_to_append and auto_log_keluar:
            self.log_gui(f"\n[INFO] Mengirim data stok keluar ke LOG_KELUAR...", "info")
            ws_log.append_rows(logs_keluar_to_append)
        elif not auto_log_keluar:
            self.log_gui(f"\n[INFO] LOG_KELUAR DILEWATI sesuai opsi pra-eksekusi (mode manual).", "kuning")

        # WRITE LOG EXCEL LOCAL
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        if success_logs:
            wb_success = Workbook()
            ws_success = wb_success.active
            ws_success.append(["ID Tugas", "ID Desain", "Produksi (Pcs/Lbr)", "Keterangan"])
            for lg in success_logs: ws_success.append(lg)
            wb_success.save(os.path.join(LOG_BERHASIL_DIR, f"berhasil_{timestamp}.xlsx"))

        if fail_logs:
            wb_fail = Workbook()
            ws_fail = wb_fail.active
            ws_fail.append(["SKU", "Jumlah Target", "Keterangan"])
            for lg in fail_logs: ws_fail.append(lg)
            wb_fail.save(os.path.join(LOG_GAGAL_DIR, f"gagal_{timestamp}.xlsx"))

        if warnings_log:
            wb_warn = Workbook()
            ws_warn = wb_warn.active
            ws_warn.append(["SKU", "Jml", "Keterangan"])
            for lg in warnings_log: ws_warn.append(lg)
            wb_warn.save(os.path.join(LOG_PERINGATAN_DIR, f"peringatan_{timestamp}.xlsx"))

        self.progress.set(1.0)

        # Summary SKU produk non-stiker yg di-skip (kalau ada)
        sheet_stats = non_stiker_stats["sheet"]
        gk_stats = non_stiker_stats["gantungan_kunci"]
        if sheet_stats["pesanan"] > 0 or gk_stats["pesanan"] > 0:
            self.log_gui("\n[INFO] Produk non-stiker yang di-skip (tidak dicetak):", "kuning")
            if sheet_stats["pesanan"] > 0:
                self.log_gui(
                    f"  • Stiker SHEET: {len(sheet_stats['unique'])} SKU unik, "
                    f"{sheet_stats['pesanan']} baris pesanan ({sheet_stats['pcs']} pcs total)",
                    "kuning"
                )
            if gk_stats["pesanan"] > 0:
                self.log_gui(
                    f"  • Gantungan KUNCI: {len(gk_stats['unique'])} SKU unik, "
                    f"{gk_stats['pesanan']} baris pesanan ({gk_stats['pcs']} pcs total)",
                    "kuning"
                )
            self.log_gui(
                "  → Detail SKU per-baris ada di log/gagal/gagal_*.xlsx (keterangan 'Skip').",
                "info"
            )

        self.log_gui("\n[SELESAI] Proses telah selesai.", "info")

if __name__ == "__main__":
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("blue")
    app = BotApp()
    app.mainloop()
