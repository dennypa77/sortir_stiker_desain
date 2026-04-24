import os
import math
import re
import shutil
from datetime import datetime
from collections import defaultdict
from time import sleep

# Untuk progress bar, Anda mungkin perlu menginstal: pip install tqdm
try:
    from tqdm import tqdm
except ImportError:
    print("Peringatan: Library 'tqdm' tidak ditemukan. Progress bar tidak akan ditampilkan.")
    print("Untuk menginstal, jalankan: pip install tqdm")
    # Buat fungsi tqdm dummy jika tidak ada
    def tqdm(iterable, **kwargs):
        return iterable

from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2.errors import PdfReadError

# ===================== KONFIGURASI =====================
MASTER_FOLDER = r"C:\Users\Stickitup123\Downloads\pdf"  # Folder berisi file PDF, mis: "1-kucing.pdf", "700-SKL-50pcsA.pdf", dst
HOT_FOLDER = r"C:\Users\Stickitup123\Downloads\hot_file\hasil"
EXCEL_FILE = r"C:\Users\Stickitup123\Downloads\hot_file\data.xlsx"

# Konfigurasi folder log, tidak perlu diubah
LOG_BERHASIL_DIR = os.path.join(HOT_FOLDER, "log", "berhasil")
LOG_GAGAL_DIR = os.path.join(HOT_FOLDER, "log", "gagal")
LOG_PERINGATAN_DIR = os.path.join(HOT_FOLDER, "log", "peringatan")
LOG_DATA_DIR = os.path.join(HOT_FOLDER, "log")

def clear_hotfolder_pdfs():
    """
    FITUR BARU: Menghapus semua file PDF di HOT_FOLDER.
    Fungsi ini secara spesifik akan mengabaikan folder 'log' dan isinya,
    sehingga riwayat log tetap aman.
    """
    print("Membersihkan HOT_FOLDER dari file PDF sebelumnya...")
    try:
        for name in os.listdir(HOT_FOLDER):
            full_path = os.path.join(HOT_FOLDER, name)
            # Lewati folder 'log'
            if os.path.isdir(full_path) and name.lower() == "log":
                continue
            # Hapus hanya file yang berakhiran .pdf
            if os.path.isfile(full_path) and name.lower().endswith('.pdf'):
                os.remove(full_path)
        print("Pembersihan selesai.")
    except Exception as e:
        print(f"\033[91mError saat membersihkan HOT_FOLDER: {e}\033[0m")


def create_file_cache(master_folder: str) -> dict:
    """Membuat cache dari semua file PDF di master folder untuk pencarian cepat."""
    print("Membuat cache daftar file dari MASTER_FOLDER...")
    cache = defaultdict(list)
    try:
        for filename in os.listdir(master_folder):
            if filename.lower().endswith(".pdf"):
                name, _ = os.path.splitext(filename)
                cache[name.lower()].append(os.path.join(master_folder, filename))
        print(f"Cache dibuat. Ditemukan {sum(len(v) for v in cache.values())} file PDF.")
        return cache
    except FileNotFoundError:
        print(f"\033[91mError: MASTER_FOLDER '{master_folder}' tidak ditemukan.\033[0m")
        return None

def extract_id_from_sku(sku: str) -> str:
    """Mengekstrak ID Desain LENGKAP dari SKU untuk keperluan logging/display."""
    pattern = re.compile(r'[-_ ]?(10|20|50|100)pcs(ab|a|b)?$', re.IGNORECASE)
    design_id = pattern.sub('', sku)
    return design_id.strip()

def extract_numeric_id_from_sku(sku: str) -> str | None:
    """Mengekstrak HANYA bagian numerik di awal SKU sebagai kunci pencarian."""
    match = re.match(r'^\d+', sku.strip())
    return match.group(0) if match else None

def get_task_variants(sku: str) -> dict:
    """Menentukan varian tugas (halaman dan pcs) dari SKU."""
    sku_lower = sku.lower()
    if sku_lower.endswith("100pcsab"):
        return {'-A': {'page': 1, 'pcs_multiplier': 50}, '-B': {'page': 3, 'pcs_multiplier': 50}}
    page = 3 if "b" in sku_lower[-6:] else 1
    suffix = '-B' if page == 3 else ('-A' if "a" in sku_lower[-6:] else '')
    pcs_match = re.search(r'(10|20|50|100)pcs', sku_lower)
    pcs = int(pcs_match.group(1)) if pcs_match else 50
    return {suffix: {'page': page, 'pcs_multiplier': pcs}}

def find_file_with_priority(file_cache: dict, numeric_id: str, warnings_log: list) -> tuple[str | None, str | None]:
    """
    Mencari file di cache berdasarkan ID numerik, dengan mekanisme pemilihan otomatis jika ambigu.
    Logika "versioptimal" telah diperbarui agar lebih fleksibel.
    """
    all_candidates = []
    # Kumpulkan semua kandidat file yang cocok dengan ID numerik
    for name_lower, paths in file_cache.items():
        if name_lower.startswith(numeric_id):
            # Memastikan kecocokan yang tepat (misal: '3' harus cocok dengan '3-nama' bukan '30-nama')
            if len(name_lower) > len(numeric_id) and not name_lower[len(numeric_id)].isalnum():
                 all_candidates.extend(paths)
            elif len(name_lower) == len(numeric_id):
                 all_candidates.extend(paths)

    if not all_candidates:
        return None, None

    # DIPERBARUI: Logika pencarian 'versioptimal' yang lebih fleksibel
    # Memeriksa apakah 'versioptimal' ada di dalam nama file (case-insensitive)
    optimal_candidates = [p for p in all_candidates if "versioptimal" in os.path.basename(p).lower()]
    standard_candidates = [p for p in all_candidates if p not in optimal_candidates]

    # Prioritaskan file optimal jika ditemukan
    if optimal_candidates:
        if len(optimal_candidates) > 1:
            warnings_log.append((f"ID: {numeric_id}", "", f"Peringatan: Ditemukan duplikat file optimal ({len(optimal_candidates)} buah). Menggunakan file terpendek."))
            optimal_candidates.sort(key=len)
        return optimal_candidates[0], 'optimal'

    # Jika tidak ada file optimal, gunakan file standar
    if len(standard_candidates) > 1:
        warnings_log.append((f"ID Numerik: {numeric_id}", "", f"Peringatan: Pencarian ambigu, ditemukan {len(standard_candidates)} file standar. Otomatis memilih nama file terpendek."))
        standard_candidates.sort(key=len)
        return standard_candidates[0], 'standard'
    
    if len(standard_candidates) == 1:
        return standard_candidates[0], 'standard'

    return None, None

def extract_pages(pdf_input_path: str, pdf_output_path: str, page_to_extract: int):
    """Mengekstrak SATU halaman spesifik."""
    try:
        with open(pdf_input_path, "rb") as infile:
            reader = PdfReader(infile)
            writer = PdfWriter()
            if page_to_extract - 1 < len(reader.pages):
                page = reader.pages[page_to_extract - 1]
                writer.add_page(page)
                with open(pdf_output_path, "wb") as outfile:
                    writer.write(outfile)
    except Exception as e:
        raise IOError(f"Gagal membuka atau memproses file PDF: {pdf_input_path}. Error: {e}")


used_filenames = defaultdict(int)
def get_next_filename(task_id: int, design_id: str, suffix: str = "") -> str:
    """Membuat nama file unik berdasarkan ID Tugas."""
    sanitized_id = re.sub(r'[\\/*?:"<>|]', "-", design_id)
    base_key = f"Tugas-{task_id:03d}-{sanitized_id}{suffix}"
    count = used_filenames[base_key]
    used_filenames[base_key] += 1
    if count == 0: return f"{base_key}.pdf"
    else: return f"{base_key} - Copy ({count}).pdf"

# --- MAIN ---
# Pastikan semua direktori yang dibutuhkan ada
os.makedirs(HOT_FOLDER, exist_ok=True)
os.makedirs(LOG_BERHASIL_DIR, exist_ok=True)
os.makedirs(LOG_GAGAL_DIR, exist_ok=True)
os.makedirs(LOG_PERINGATAN_DIR, exist_ok=True)
os.makedirs(LOG_DATA_DIR, exist_ok=True)

# PANGGIL FITUR BARU: Bersihkan HOT_FOLDER sebelum memulai
clear_hotfolder_pdfs()

if not os.path.exists(EXCEL_FILE):
    print(f"\033[91mFile Excel tidak ditemukan di: {EXCEL_FILE}\033[0m")
    exit()

file_cache = create_file_cache(MASTER_FOLDER)
if file_cache is None: exit()

timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
new_excel_file = os.path.join(LOG_DATA_DIR, f"data_{timestamp}.xlsx")
shutil.copyfile(EXCEL_FILE, new_excel_file)

print("\nTahap 1: Membaca dan menggabungkan data pesanan...")
wb = load_workbook(EXCEL_FILE)
ws = wb.active
row_start = 2 # Asumsi data dimulai dari baris ke-2
max_row = ws.max_row
aggregated_tasks = defaultdict(lambda: {'total_pieces': 0, 'original_skus': set()})
fail_logs = []

for row_idx in range(row_start, max_row + 1):
    sku_val, jml_val = ws.cell(row_idx, 1).value, ws.cell(row_idx, 2).value
    
    if not all([sku_val, jml_val]):
        fail_logs.append((str(sku_val or ''), str(jml_val or ''), f"Gagal - Data baris {row_idx} tidak lengkap (SKU atau Jumlah kosong)"))
        continue
    try:
        original_jml = int(jml_val)
        original_sku = str(sku_val).strip()
        
        numeric_id = extract_numeric_id_from_sku(original_sku)
        if not numeric_id:
            fail_logs.append((original_sku, str(original_jml), "Gagal - Tidak dapat menemukan ID Numerik di awal SKU."))
            continue
        
        design_id_display = extract_id_from_sku(original_sku)
        variants = get_task_variants(original_sku)
        for suffix, variant_details in variants.items():
            task_key = (numeric_id, design_id_display, suffix)
            pieces_to_add = variant_details['pcs_multiplier'] * original_jml
            aggregated_tasks[task_key]['total_pieces'] += pieces_to_add
            aggregated_tasks[task_key]['original_skus'].add(original_sku)
    except (ValueError, TypeError):
        fail_logs.append((str(sku_val), str(jml_val), f"Gagal - 'Jumlah' di baris {row_idx} harus berupa angka"))
        continue

print(f"Agregasi selesai. Ditemukan {len(aggregated_tasks)} tugas cetak unik.")

# --- TAHAP 2: PEMROSESAN ---
print("\nTahap 2: Memproses tugas cetak...")
success_logs, warnings, total_prints_needed, task_id_counter = [], [], 0, 0

progress_bar_settings = {'total': len(aggregated_tasks), 'desc': "Memproses Tugas", 'unit': "tugas", 'ncols': 100, 'bar_format': "{desc}: {percentage:3.0f}%|\033[93m{bar}\033[0m| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]"}

with tqdm(**progress_bar_settings) as progress_bar:
    for task_key, task_data in aggregated_tasks.items():
        task_id_counter += 1
        numeric_id, design_id_display, suffix = task_key
        total_pieces = task_data['total_pieces']
        
        progress_bar.set_description(f"Tugas {task_id_counter}: ID {numeric_id}")

        if total_pieces <= 0:
            keterangan = "Sukses - Tidak ada yang perlu dicetak (Jumlah total 0)"
            success_logs.append((f"Tugas-{task_id_counter:03d}", design_id_display, str(total_pieces), keterangan))
            progress_bar.update(1)
            continue

        found_file, version_type = find_file_with_priority(file_cache, numeric_id, warnings)
        if not found_file:
            fail_logs.append((f"Agregat untuk: {design_id_display}", str(total_pieces), "Gagal - File desain tidak ditemukan (ID Numerik tidak ditemukan di Master Folder)"))
            progress_bar.update(1)
            continue

        batch_size_per_page = 100.0 if version_type == 'optimal' else 50.0
        page_to_extract = 3 if suffix == '-B' else 1
        num_pages_to_print = int(math.ceil(total_pieces / batch_size_per_page))

        try:
            # DIPERBARUI: Tambahkan akhiran khusus untuk versi optimal
            final_suffix = suffix
            if version_type == 'optimal':
                final_suffix += "-VERSIOPTIMAL"

            for _ in range(num_pages_to_print):
                # Gunakan akhiran yang sudah dimodifikasi untuk membuat nama file
                out_filename = get_next_filename(task_id_counter, design_id_display, final_suffix)
                out_path = os.path.join(HOT_FOLDER, out_filename)
                extract_pages(found_file, out_path, page_to_extract)
            
            total_prints_needed += num_pages_to_print
            keterangan = f"Sukses - Cetak {num_pages_to_print} lbr (File: {version_type})"
            success_logs.append((f"Tugas-{task_id_counter:03d}", design_id_display, str(total_pieces), keterangan))
        except (PdfReadError, IOError) as e:
            fail_logs.append((f"Agregat untuk: {design_id_display}", str(total_pieces), f"Gagal - {e}"))
        
        progress_bar.update(1)
        sleep(0.01)

    progress_bar.bar_format = "{desc}: {percentage:3.0f}%|\032[92m{bar}\033[0m| {n_fmt}/{total_fmt} [{elapsed}]"
    progress_bar.set_description("Proses Selesai ✔")
    progress_bar.refresh()

# --- SELESAI & LOGGING ---
print(f"\n\n\033[92mBerhasil diproses. Total cetak {total_prints_needed} lembar dari {len(aggregated_tasks)} tugas.\033[0m")

if success_logs:
    wb_success = Workbook()
    ws_success = wb_success.active
    ws_success.title = "Log Berhasil (Agregat Total)"
    ws_success.append(["ID Tugas", "ID Desain (dari SKU)", "Total Pcs", "Keterangan"])
    for log_entry in success_logs:
        ws_success.append(log_entry)
    berhasil_path = os.path.join(LOG_BERHASIL_DIR, f"berhasil_agregat_{timestamp}.xlsx")
    wb_success.save(berhasil_path)
    print(f"Log berhasil disimpan di: {berhasil_path}")

if fail_logs:
    wb_fail = Workbook()
    ws_fail = wb_fail.active
    ws_fail.title = "Log Gagal"
    ws_fail.append(["SKU/ID Desain", "Jumlah/Total Pcs", "Keterangan"])
    for log_entry in fail_logs:
        ws_fail.append(log_entry)
    gagal_path = os.path.join(LOG_GAGAL_DIR, f"gagal_{timestamp}.xlsx")
    wb_fail.save(gagal_path)
    print(f"\033[91mAda {len(fail_logs)} item yang GAGAL diproses.\033[0m")
    print(f"Log gagal disimpan di: {gagal_path}")

if warnings:
    wb_warn = Workbook()
    ws_warn = wb_warn.active
    ws_warn.title = "Log Peringatan"
    ws_warn.append(["SKU/ID Desain", "Jumlah/Total Pcs", "Keterangan"])
    for log_entry in warnings:
        ws_warn.append(log_entry)
    peringatan_path = os.path.join(LOG_PERINGATAN_DIR, f"peringatan_{timestamp}.xlsx")
    wb_warn.save(peringatan_path)
    print(f"\033[93mAda {len(warnings)} PERINGATAN sistem yang perlu diperhatikan.\033[0m")
    print(f"Log peringatan disimpan di: {peringatan_path}")

print(f"Data Excel asli yang diproses disimpan di: {new_excel_file}")
